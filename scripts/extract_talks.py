#!/usr/bin/env python3
"""Extract talk records from the organizer's xlsx into public data/talks/*.yaml.

Source xlsx stays private (OneDrive, never committed). This script is the
repeatable step: rerun after every xlsx edit, then commit the regenerated yaml.

Usage:
    python3 scripts/extract_talks.py [--source PATH] [--sheets 2026,2026-online]

By default every sheet named like a year ("2027") or year-online ("2027-online")
is processed automatically — adding a new year tab in the xlsx needs no script
change. data/talks/all.yaml is always rebuilt from ALL per-sheet yaml files on
disk, so partial --sheets runs never drop other years' talks from the site.
"""
import argparse
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
import yaml

DEFAULT_SOURCE = (
    "/Users/ccenedese/Library/CloudStorage/OneDrive-DelftUniversityofTechnology/"
    "TU Delft/00.Doc/05.DCSC/03.lunch colloquium & poster/06.Website/DCSC_colloquia/"
    "01.Colloquia organization/Lunch colloquia version_v1.xlsx"
)
# Sheets are auto-discovered by this pattern; pass --sheets to override.
SHEET_NAME_RE = re.compile(r"^\d{4}(-online)?$")
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data" / "talks"
SLIDES_DIR = Path(__file__).resolve().parent.parent / "static" / "slides"
SLIDES_EXTENSIONS = [".pdf", ".pptx", ".ppt"]
AVATARS_DIR = Path(__file__).resolve().parent.parent / "static" / "avatars"
AVATAR_EXTENSIONS = [".jpg", ".jpeg", ".png", ".webp"]

HEADER = [
    "date", "location", "time", "speaker", "gp", "role", "affiliation",
    "status_raw", "title", "abstract", "tags_raw", "bio", "pic", "video",
    "notes", "youtube_url",
]

YT_ID_RE = re.compile(r"(?:youtu\.be/|v=)([A-Za-z0-9_-]{6,})")

# Raw xlsx affiliation string -> (short, full-with-department).
# Curated by hand since the xlsx spelling is inconsistent and department info
# isn't always present in the same column (sometimes folded into the string).
AFFILIATION_MAP = {
    "University of Campinas": ("UNICAMP", "University of Campinas"),
    "TUD(DCSC)": ("TU Delft", "TU Delft — Delft Center for Systems and Control (DCSC)"),
    "TUD (DCSC)": ("TU Delft", "TU Delft — Delft Center for Systems and Control (DCSC)"),
    "TU Delft -- Department of Software Technology": ("TU Delft", "TU Delft — Department of Software Technology"),
    "Denmark Technical University / DTU Compute / Cognitive systems": (
        "DTU", "Technical University of Denmark — DTU Compute, Cognitive Systems",
    ),
    "ETH": ("ETH Zurich", "ETH Zurich — Automatic Control Laboratory (IfA)"),
    "UBC": ("UBC", "University of British Columbia — Department of Mechanical Engineering"),
    "Electrical & Computer Eng. at University of Minnesota": (
        "UMN", "University of Minnesota — Department of Electrical and Computer Engineering",
    ),
    "KTH": ("KTH", "KTH Royal Institute of Technology — Department of Decision and Control Systems"),
    "KU Leuven": ("KU Leuven", "KU Leuven — Department of Electrical Engineering (ESAT-STADIUS)"),
    "AI4I": ("AI4I Institute", "AI4I Institute — RIAS Lab, Turin, Italy"),
    "ESTACA (École Supérieure des Techniques Aéronautiques et de Construction Automobile)": (
        "ESTACA", "ESTACA — École Supérieure des Techniques Aéronautiques et de Construction Automobile, France",
    ),
    "Caltech": ("Caltech", "California Institute of Technology — Department of Computing and Mathematical Sciences"),
    "University of Opole": ("UO", "University of Opole — Institute of Computer Science"),
    "University of New Mexico (USA)": ("UNM", "University of New Mexico — Department of Mechanical Engineering"),
    "Massachusetts Institute of Technology (USA)": ("MIT", "Massachusetts Institute of Technology — Department of Chemical Engineering"),
    "Norwegian University of Science and Technology (NOR)": (
        "NTNU", "Norwegian University of Science and Technology — Department of Engineering Cybernetics",
    ),
    "ETH Zurich (CH)": ("ETH Zurich", "ETH Zurich — Automatic Control Laboratory"),
    "Max Planck Institute for Intelligent Systems (Germany)": (
        "MPI-IS", "Max Planck Institute for Intelligent Systems — Learning and Dynamical Systems Group",
    ),
    "University of Michigan (USA)": ("U-M", "University of Michigan — Electrical Engineering and Computer Science Department"),
    "University of Brescia (Italy)": ("UniBS", "University of Brescia — Dept. of Information Engineering"),
    "Cornell University (USA)": ("Cornell", "Cornell University — Sibley School of Mechanical and Aerospace Engineering"),
    "TUD (Chemical engineering)": ("TU Delft", "TU Delft — Department of Chemical Engineering"),
    "TU Delft Computer Science": ("TU Delft", "TU Delft — Faculty of Electrical Engineering, Mathematics and Computer Science"),
    "TU Eindhoven": ("TU/e", "Eindhoven University of Technology"),
    "Politecnico di Milano": ("PoliMi", "Politecnico di Milano"),
    "Linköping University": ("LiU", "Linköping University"),
    "Leiden University": ("Leiden", "Leiden University"),
    "LUMC": ("LUMC", "Leiden University Medical Center"),
    "ETH Zürich": ("ETH Zurich", "ETH Zurich"),
    "Dep. of Biomedical Engineering, HAW Hamburg": ("HAW Hamburg", "HAW Hamburg — Department of Biomedical Engineering"),
}

# The only 4 research clusters the site is allowed to show. Everything from
# the xlsx gets folded into one of these or dropped (never invented).
TAG_CANON = {
    "control and learning": "Control and Learning",
    "modeling and system identification": "Modeling and System Identification",
    "modelling and system identification": "Modeling and System Identification",
    "optimization": "Optimization",
    "and optimization": "Optimization",  # bad comma-split fragment in source
    "systems and signal analysis": "Systems and Signal Analysis",
}


def resolve_affiliation(raw):
    raw = clean(raw)
    if not raw:
        return "TBD", "TBD"
    mapped = AFFILIATION_MAP.get(raw)
    if mapped:
        return mapped
    print(f"warn: no affiliation mapping for {raw!r} — using raw string for both", file=sys.stderr)
    return raw, raw


def canonicalize_tags(raw_tags):
    canon = []
    for raw in raw_tags:
        key = raw.strip().lower()
        mapped = TAG_CANON.get(key)
        if mapped:
            if mapped not in canon:
                canon.append(mapped)
        else:
            print(f"warn: dropping non-canonical tag {raw!r}", file=sys.stderr)
    return canon


# Raw xlsx role string -> normalized abbreviation shown on the site.
ROLE_CANON = {
    "ap": "Asst. Prof.",
    "assistant professor": "Asst. Prof.",
    "assistant prof": "Asst. Prof.",
    "associate prof": "Associate Prof.",
    "associate professor": "Associate Prof.",
    "professor": "Professor",
    "postdoc": "Postdoc",
    "phd": "PhD",
    "msc": "MSc",
    "research group leader": "Research Group Leader",
}


def resolve_role(raw):
    raw = clean(raw)
    if not raw:
        return "TBD"
    mapped = ROLE_CANON.get(raw.strip().lower())
    if mapped:
        return mapped
    print(f"warn: no role mapping for {raw!r} — using raw string", file=sys.stderr)
    return raw


def slugify(text):
    text = re.sub(r"[^\w\s-]", "", text or "").strip().lower()
    return re.sub(r"[\s_]+", "-", text)


def youtube_id(url):
    if not url:
        return None
    match = YT_ID_RE.search(url)
    return match.group(1) if match else None


def clean(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def clean_or_tbd(value):
    return clean(value) or "TBD"


def find_slides(talk_id):
    """Convention over configuration: drop static/slides/<talk-id>.pdf (or .pptx/.ppt)
    to attach slides — no yaml editing needed, this just picks it up on next run."""
    for ext in SLIDES_EXTENSIONS:
        candidate = SLIDES_DIR / f"{talk_id}{ext}"
        if candidate.exists():
            return f"/slides/{candidate.name}"
    return ""


def find_avatar(speaker):
    """Drop static/avatars/<speaker-slug>.jpg (or .jpeg/.png/.webp) to attach a photo.
    Keyed by speaker, not talk id, so one photo covers every talk by that person."""
    slug = slugify(speaker)
    if not slug:
        return ""
    for ext in AVATAR_EXTENSIONS:
        candidate = AVATARS_DIR / f"{slug}{ext}"
        if candidate.exists():
            return f"/avatars/{candidate.name}"
    return ""


def parse_sheet(ws, mode):
    talks = []
    seen_ids = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        cells = row[1:1 + len(HEADER)]
        record = dict(zip(HEADER, cells))
        if not record.get("date"):
            continue  # no date at all: not a real slot
        if not clean(record.get("speaker")):
            continue  # date-only placeholder: blank/skipped week, not a real slot

        talk_date = record["date"].date() if hasattr(record["date"], "date") else record["date"]
        raw_tags = [t.strip() for t in re.split(r"[,;]", record.get("tags_raw") or "") if t.strip()]
        tags = canonicalize_tags(raw_tags)
        yt_url = clean(record.get("youtube_url"))
        affiliation_short, affiliation_full = resolve_affiliation(record.get("affiliation"))

        base_id = f"{talk_date.isoformat()}-{slugify(record.get('speaker')) or 'tbd'}"
        talk_id = base_id
        suffix = 2
        while talk_id in seen_ids:
            talk_id = f"{base_id}-{suffix}"
            suffix += 1
        seen_ids.add(talk_id)

        talks.append({
            "id": talk_id,
            "date": talk_date.isoformat(),
            "time": clean(record.get("time")),
            "mode": mode,
            "location": clean_or_tbd(record.get("location")),
            "speaker": clean_or_tbd(record.get("speaker")),
            "affiliation_short": affiliation_short,
            "affiliation_full": affiliation_full,
            "role": resolve_role(record.get("role")),
            "title": clean_or_tbd(record.get("title")),
            "abstract": clean_or_tbd(record.get("abstract")),
            "bio": clean_or_tbd(record.get("bio")),
            "tags": tags,
            "youtube_url": yt_url,
            "youtube_id": youtube_id(yt_url),
            "status": "past" if talk_date <= date.today() else "upcoming",
            "notes": clean(record.get("notes")),
            "links": {"scholar": "TBD", "homepage": "", "verified": False},
            "slides_url": find_slides(talk_id),
            "avatar_url": find_avatar(record.get("speaker")),
        })
    return talks


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=DEFAULT_SOURCE)
    parser.add_argument("--sheets", default="",
                        help="comma-separated sheet names; default: every sheet matching YYYY or YYYY-online")
    args = parser.parse_args()

    source = Path(args.source)
    if not source.exists():
        sys.exit(f"xlsx not found: {source}")

    wb = openpyxl.load_workbook(source, data_only=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.sheets:
        sheet_names = [s.strip() for s in args.sheets.split(",") if s.strip()]
    else:
        sheet_names = [s for s in wb.sheetnames if SHEET_NAME_RE.match(s)]
        if not sheet_names:
            sys.exit("no sheets matching YYYY or YYYY-online found; pass --sheets explicitly")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"skip: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        mode = "online" if "online" in sheet_name else "in-person"
        talks = parse_sheet(wb[sheet_name], mode)
        out_path = OUTPUT_DIR / f"{sheet_name}.yaml"

        existing = {}
        if out_path.exists():
            prior = yaml.safe_load(out_path.read_text()) or {}
            existing = {t["id"]: t for t in prior.get("talks", [])}

        for talk in talks:
            prior_talk = existing.get(talk["id"])
            if prior_talk and prior_talk.get("links"):
                talk["links"] = prior_talk["links"]

        out_path.write_text(yaml.safe_dump(
            {"talks": talks}, sort_keys=False, allow_unicode=True, width=100,
        ))
        print(f"wrote {out_path} ({len(talks)} talks)")

    # Rebuild all.yaml from EVERY per-sheet file on disk (not just this run's
    # sheets) so a partial --sheets run can never drop other years from the site.
    all_talks = []
    for path in sorted(OUTPUT_DIR.glob("*.yaml")):
        if path.name == "all.yaml":
            continue
        data = yaml.safe_load(path.read_text()) or {}
        all_talks.extend(data.get("talks", []))

    all_path = OUTPUT_DIR / "all.yaml"
    all_path.write_text(yaml.safe_dump(
        {"talks": all_talks}, sort_keys=False, allow_unicode=True, width=100,
    ))
    print(f"wrote {all_path} ({len(all_talks)} talks, combined)")


if __name__ == "__main__":
    main()
